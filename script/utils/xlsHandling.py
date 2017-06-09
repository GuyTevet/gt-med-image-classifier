#!/usr/bin/env python

"""
    File name:          smartCrop.py
    Author:             Guy Tevet
    Date created:       9/6/2017
    Date last modified: 9/6/2017
    Description:        xls utility functions for parsing and 
                        editing experiments.xls
"""

from experimentDefines import *
from openpyxl import Workbook , load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill , PatternFill , Border , Side
from time import localtime , strftime
import os
from copy import copy

def loadExperiments():
        wb = load_workbook(filename=EXPERIMENTS_FILE)
        ws = wb[EXPERIMENTS_TAB]
        return [wb,ws] 

def saveExperiments(wb):
        wb.save(EXPERIMENTS_FILE)
        return

def getExperimentRow(experimentNum):
        [wb,ws] = loadExperiments()
        for row in range(1,ws.max_row+1):
                cell = ws[EXPERIMENT_NUM_COL + str(row)]
                if(cell.value == experimentNum):
                        return str(row)
        raise Exception('row not found')
        
def getAttributeValue(attribute,experimentNum):
        [wb,ws] = loadExperiments()
        for col in range(1,ws.max_column+1):
                coll_letter = get_column_letter(col)
                cell = ws[coll_letter + ATTRIBUTES_ROW]
                if(cell.value == attribute):
                        return ws[coll_letter + getExperimentRow(experimentNum)].value
        raise Exception('attribute not found')

def getAttributeCol(attribute):
        [wb,ws] = loadExperiments()
        for col in range(1,ws.max_column+1):
                coll_letter = get_column_letter(col)
                cell = ws[coll_letter + ATTRIBUTES_ROW]
                if(cell.value == attribute):
                        return coll_letter
        raise Exception('attribute not found')

def copyRow(src_file,src_row,dest_file,dest_row,src_tab = EXPERIMENTS_TAB,dest_tab = RESULTS_TAB):
        #open files
        src_wb = load_workbook(filename=src_file)
        src_ws = src_wb[src_tab]
        dest_wb = load_workbook(filename=dest_file)
        dest_ws = dest_wb[dest_tab]
        
        #copy row
        for col in range(1,src_ws.max_column+1):
                coll_letter = get_column_letter(col)
                dest_ws[coll_letter + str(dest_row)].value =\
                copy(src_ws[coll_letter + str(src_row)].value)
                dest_ws[coll_letter + str(dest_row)].fill =\
                copy(src_ws[coll_letter + str(src_row)].fill)
                dest_ws[coll_letter + str(dest_row)].border =\
                copy(src_ws[coll_letter + str(src_row)].border)
                dest_ws[coll_letter + str(dest_row)].font =\
                copy(src_ws[coll_letter + str(src_row)].font)

        #save dest file
        dest_wb.save(dest_file)


        

